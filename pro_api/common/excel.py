import openpyxl
import json
from common.os_path import case_dir # 引入测试用例的路径
from common.request import Http_Request

class Case:

    def __init__(self):  # 初始化测试数据 访问接口只需要用到：url data  header，初始值为None，读取EXCEL中对应的值，再赋上
        self.case_id = None
        self.url = None
        self.data = None
        self.header = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None



class DoExcel:


    def __init__(self,file_name):
        try:
            self.workbook = openpyxl.load_workbook(filename=file_name) # 使用python中openpyxl打开excel
        except FileNotFoundError as e:
            print("Excel不存在:{}".format(e))
            raise e


    def get_case(self,sheet_name):
        sheet = self.workbook[sheet_name] # 根据打开的excel 定位excel中的表单sheet
        max_row = sheet.max_row # 获取sheet表单中的最大行

        test_data = [] # 定义一个空的list装载用例
        for i in range(2,max_row+1):
            case =Case() # 实例化上面的Case，下面要读取excel中的测试数据，赋值给初始None
            case.case_id = sheet.cell(row = i ,column = 1).value # 读取excel中case_id
            case.url = sheet.cell(row = i ,column = 2).value # 读取excel中url
            case.data = sheet.cell(row = i ,column = 3).value # 读取excel中data
            case.header = sheet.cell(row = i ,column = 4).value # 读取excel中header
            case.title = sheet.cell(row = i ,column = 5).value # 读取excel中title
            case.method = sheet.cell(row = i ,column = 6).value # 读取excel中请求方式
            case.expected = sheet.cell(row = i ,column = 7).value # 读取excel中期望结果
            test_data.append(case) # 读取出来的测试数据放到list

        return test_data


if __name__ == '__main__':
        case_data = DoExcel(case_dir)
        case = case_data.get_case('In_applying_for')
        for case_run in case:  # 遍历出组装好的测试用例
            print('case信息:{}'.format(case_run.__dict__))

            # 用读取出来的测试数据，发起requests请求
            # resp = Http_Request(case_run.method,url=case_run.url,headers=eval(case_run.header))
            # resp_dict = resp.get_json()
            # print("用例标题:{0},请求结果:{1}".format(case_run.title,resp_dict))
            #
            # if resp_dict['code'] == '000000':
            #     print("PASS")
            #
            # else:
            #     print("FAILED")







